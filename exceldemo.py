import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait


Dict={}


pathforuploadfile="C:\\Users\\SHWETA\\Downloads\\download.xlsx"
driver=webdriver.Chrome()
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.implicitly_wait(10)
driver.find_element(By.ID,"downloadButton").click()
#edit excel with updated value

book=openpyxl.load_workbook(pathforuploadfile)
sheet=book.active
for cl in range(1,sheet.max_column+1):
    if sheet.cell(row=1, column=cl).value =="price":
        Dict["col"]=cl
for rw in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        if sheet.cell(row=rw, column=j).value =="Apple":
            Dict["row"]=rw

sheet.cell(row= Dict["row"], column=Dict["col"]).value == "876"
book.save(pathforuploadfile)

#upload file
driver.find_element(By.CSS_SELECTOR,"input[type='file']").send_keys(pathforuploadfile)
wait=WebDriverWait(driver,10)

wait.until(expected_conditions.visibility_of_element_located((By.XPATH,"//div[@class='Toastify']")))
print(driver.find_element(By.XPATH,"//div[@class='Toastify']").text)










